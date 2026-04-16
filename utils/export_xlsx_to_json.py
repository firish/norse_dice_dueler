"""
export_xlsx_to_json.py
----------------------
Reads Fjold_Master_Design_v1.1.xlsx (or any version) and exports each
game-content tab as a JSON file into /data/.

Usage:
    # Export everything
    python3 utils/export_xlsx_to_json.py

    # Export specific sheets
    python3 utils/export_xlsx_to_json.py dice_types god_powers runes

    # Point at a different workbook
    python3 utils/export_xlsx_to_json.py --xlsx path/to/file.xlsx

Adding a new sheet:
    1. Write a parse_<snake_case_sheet_name>(ws) function below.
    2. Add an entry to SHEET_EXPORTERS: {"Sheet Name": (parse_fn, "output_filename.json")}.
    3. Run the script - the new file appears in /data/.
"""

from __future__ import annotations

import argparse
import json
import math
import pathlib
import sys
from typing import Any, Callable

import openpyxl

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
REPO_ROOT = pathlib.Path(__file__).resolve().parent.parent
DATA_DIR = REPO_ROOT / "data"

_DEFAULT_XLSX_GLOB = "Fjold_Master_Design_*.xlsx"


def _find_xlsx() -> pathlib.Path:
    matches = sorted(REPO_ROOT.glob(_DEFAULT_XLSX_GLOB))
    if not matches:
        sys.exit(f"[error] No workbook matching '{_DEFAULT_XLSX_GLOB}' found in {REPO_ROOT}")
    if len(matches) > 1:
        print(f"[warn] Multiple workbooks found - using newest: {matches[-1].name}")
    return matches[-1]


# ---------------------------------------------------------------------------
# Generic helpers
# ---------------------------------------------------------------------------

def _rows(ws, skip_header_rows: int = 0) -> list[tuple]:
    """Return all non-empty rows from a worksheet, skipping the first N rows."""
    result = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < skip_header_rows:
            continue
        if any(v is not None for v in row):
            result.append(row)
    return result


def _find_header_row(ws, first_col_value: str) -> int:
    """Return the 0-based index of the row where column A == first_col_value."""
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if row and str(row[0]).strip() == first_col_value:
            return i
    raise ValueError(f"Header row '{first_col_value}' not found in sheet '{ws.title}'")


def _clean(value: Any) -> Any:
    """Convert Excel artefacts (formula strings, NaN, etc.) to clean Python values."""
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    if isinstance(value, str):
        stripped = value.strip()
        return stripped if stripped else None
    return value


def _power_budget(axe, arrow, helmet, shield, hand, bordered) -> float:
    """Constitution C04 formula."""
    return round(
        (axe or 0) * 1.0
        + (arrow or 0) * 1.0
        + (helmet or 0) * 1.0
        + (shield or 0) * 1.0
        + (hand or 0) * 0.8
        + (bordered or 0) * 1.2,
        2,
    )


# ---------------------------------------------------------------------------
# Sheet parsers
# ---------------------------------------------------------------------------

def parse_die_faces(ws) -> list[dict]:
    header_idx = _find_header_row(ws, "Face ID")
    records = []
    for row in ws.iter_rows(min_row=header_idx + 2, values_only=True):
        face_id = _clean(row[0])
        if face_id is None or face_id.startswith("VALIDATION") or face_id.startswith("Face"):
            continue
        records.append({
            "id": face_id,
            "display_name": _clean(row[1]),
            "symbol": _clean(row[2]),
            "effect": _clean(row[3]),
            "power_value": _clean(row[4]),
            "availability": _clean(row[5]),
        })
    return records


def parse_dice_types(ws) -> list[dict]:
    header_idx = _find_header_row(ws, "Die ID")
    records = []
    for row in ws.iter_rows(min_row=header_idx + 2, values_only=True):
        die_id = _clean(row[0])
        if die_id is None or die_id in ("VALIDATION", "Face Count Check"):
            continue
        axe, arrow, helmet, shield, hand, bordered = (
            int(row[3] or 0), int(row[4] or 0), int(row[5] or 0),
            int(row[6] or 0), int(row[7] or 0), int(row[8] or 0),
        )
        records.append({
            "id": die_id,
            "display_name": _clean(row[1]),
            "role": _clean(row[2]),
            "faces": {
                "axe": axe,
                "arrow": arrow,
                "helmet": helmet,
                "shield": shield,
                "hand": hand,
                "bordered_hand": bordered,
            },
            "power_budget": _power_budget(axe, arrow, helmet, shield, hand, bordered),
            "unlock": _clean(row[10]),
        })
    return records


def parse_god_powers(ws) -> list[dict]:
    header_idx = _find_header_row(ws, "Power ID")
    powers: list[dict] = []
    current: dict | None = None

    for row in ws.iter_rows(min_row=header_idx + 2, values_only=True):
        power_id = _clean(row[0])
        tier = _clean(row[3])
        if tier not in ("T1", "T2", "T3"):
            continue

        if power_id:  # start of a new power block
            current = {
                "id": power_id,
                "display_name": _clean(row[1]),
                "category": _clean(row[2]),
                "tiers": [],
                "design_notes": _clean(row[8]),
            }
            powers.append(current)

        if current is None:
            continue

        raw_dmg = _clean(row[6])
        damage: int | float | None = None
        if isinstance(raw_dmg, (int, float)):
            damage = raw_dmg
        elif isinstance(raw_dmg, str) and raw_dmg not in ("-", "-", ""):
            try:
                damage = float(raw_dmg)
            except ValueError:
                damage = None

        current["tiers"].append({
            "tier": tier,
            "cost": _clean(row[4]),
            "effect": _clean(row[5]),
            "damage": damage,
        })

    return powers


def parse_runes(ws) -> list[dict]:
    header_idx = _find_header_row(ws, "Rune ID")
    records = []
    for row in ws.iter_rows(min_row=header_idx + 2, values_only=True):
        rune_id = _clean(row[0])
        if rune_id is None:
            continue
        records.append({
            "id": rune_id,
            "display_name": _clean(row[1]),
            "category": _clean(row[2]),
            "effect": _clean(row[3]),
            "tradeoff": _clean(row[4]),
            "unlock_source": _clean(row[5]),
        })
    return records


def parse_conditions(ws) -> list[dict]:
    header_idx = _find_header_row(ws, "Condition ID")
    records = []
    for row in ws.iter_rows(min_row=header_idx + 2, values_only=True):
        cond_id = _clean(row[0])
        if cond_id is None:
            continue
        records.append({
            "id": cond_id,
            "display_name": _clean(row[1]),
            "effect": _clean(row[2]),
            "archetype_skew": _clean(row[3]),
            "rarity": _clean(row[4]),
        })
    return records


def parse_archetypes(ws) -> dict:
    # Archetype definitions
    header_idx = _find_header_row(ws, "Archetype")
    archetypes = []
    for row in ws.iter_rows(min_row=header_idx + 2, values_only=True):
        name = _clean(row[0])
        if name is None or name.startswith("vs") or name.startswith("Rock"):
            break
        archetypes.append({
            "name": name,
            "win_condition": _clean(row[1]),
            "dice_loadout": _clean(row[2]),
            "god_powers": _clean(row[3]),
            "runes": _clean(row[4]),
            "plays_against": _clean(row[5]),
        })

    # Win-rate matrix - find the block starting with "vs ->"
    matrix_rows = []
    in_matrix = False
    row_labels = []
    col_labels = []
    for row in ws.iter_rows(values_only=True):
        first = _clean(row[0])
        if first == "vs ->":
            in_matrix = True
            col_labels = [_clean(c) for c in row[1:] if _clean(c)]
            continue
        if in_matrix:
            if _clean(row[0]) is None:
                break
            label = _clean(row[0])
            values = [_clean(row[i + 1]) for i in range(len(col_labels))]
            matrix_rows.append({"archetype": label, "win_rates": dict(zip(col_labels, values))})

    return {
        "archetypes": archetypes,
        "target_win_rate_matrix": matrix_rows,
    }


def parse_balance_targets(ws) -> list[dict]:
    header_idx = _find_header_row(ws, "Metric")
    records = []
    for row in ws.iter_rows(min_row=header_idx + 2, values_only=True):
        metric = _clean(row[0])
        if metric is None:
            continue
        records.append({
            "metric": metric,
            "green": _clean(row[1]),
            "yellow": _clean(row[2]),
            "red": _clean(row[3]),
            "measurement": _clean(row[4]),
        })
    return records


def parse_pve_campaign(ws) -> dict:
    # Realms
    realm_header = _find_header_row(ws, "Order")
    realms = []
    for row in ws.iter_rows(min_row=realm_header + 2, values_only=True):
        order = _clean(row[0])
        if order is None or not str(order).isdigit():
            break
        realms.append({
            "order": int(order),
            "realm": _clean(row[1]),
            "theme": _clean(row[2]),
            "boss": _clean(row[3]),
            "unlock_reward": _clean(row[4]),
        })

    # Node structure
    node_header = _find_header_row(ws, "Node")
    nodes = []
    for row in ws.iter_rows(min_row=node_header + 2, values_only=True):
        node = _clean(row[0])
        if node is None or str(node).strip().startswith("Ascension"):
            break
        nodes.append({
            "node": str(node),
            "type": _clean(row[1]),
            "frequency": _clean(row[2]),
            "reward": _clean(row[3]),
            "notes": _clean(row[4]),
        })

    # Ascension levels
    asc_header = _find_header_row(ws, "Ascension")
    ascensions = []
    for row in ws.iter_rows(min_row=asc_header + 2, values_only=True):
        asc = _clean(row[0])
        if asc is None:
            continue
        ascensions.append({
            "level": str(asc),
            "modifier": _clean(row[1]),
            "impact": _clean(row[2]),
        })

    return {
        "realms": realms,
        "node_structure": nodes,
        "ascension_levels": ascensions,
    }


def parse_gear(ws) -> list[dict]:
    header_idx = _find_header_row(ws, "Gear ID")
    records = []
    for row in ws.iter_rows(min_row=header_idx + 2, values_only=True):
        gear_id = _clean(row[0])
        if gear_id is None:
            continue
        records.append({
            "id": gear_id,
            "display_name": _clean(row[1]),
            "slot": _clean(row[2]),
            "effect": _clean(row[3]),
            "source": _clean(row[4]),
        })
    return records


def parse_progression(ws) -> dict:
    header_idx = _find_header_row(ws, "Branch")
    branches: dict[str, list[dict]] = {}
    for row in ws.iter_rows(min_row=header_idx + 2, values_only=True):
        branch = _clean(row[0])
        node = _clean(row[1])
        if branch is None and node is None:
            continue
        if branch and branch.startswith("TOTAL"):
            continue
        if branch:
            current_branch = branch
            if current_branch not in branches:
                branches[current_branch] = []
        branches[current_branch].append({
            "node": int(node) if isinstance(node, (int, float)) else node,
            "name": _clean(row[2]),
            "effect": _clean(row[3]),
            "sp_cost": int(row[4]) if isinstance(row[4], (int, float)) else _clean(row[4]),
        })
    return branches


# ---------------------------------------------------------------------------
# Export registry
# Map: "Excel Sheet Name" -> (parser_function, "output_filename.json")
# To add a new export: add an entry here and write a parse_* function above.
# ---------------------------------------------------------------------------

SHEET_EXPORTERS: dict[str, tuple[Callable, str]] = {
    "Die Faces":             (parse_die_faces,       "die_faces.json"),
    "Dice Types":            (parse_dice_types,       "dice_types.json"),
    "God Powers":            (parse_god_powers,       "god_powers.json"),
    "Runes":                 (parse_runes,            "runes.json"),
    "Battlefield Conditions":(parse_conditions,       "conditions.json"),
    "Archetypes":            (parse_archetypes,       "archetypes.json"),
    "Balance Targets":       (parse_balance_targets,  "balance_targets.json"),
    "PvE Campaign":          (parse_pve_campaign,     "pve_campaign.json"),
    "Gear":                  (parse_gear,             "gear.json"),
    "Progression":           (parse_progression,      "progression.json"),
}

# Short aliases for CLI (lowercase, underscores)
_ALIASES = {out.replace(".json", ""): sheet for sheet, (_, out) in SHEET_EXPORTERS.items()}


# ---------------------------------------------------------------------------
# Core export logic
# ---------------------------------------------------------------------------

def export_sheets(xlsx_path: pathlib.Path, sheet_names: list[str] | None = None) -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    targets = sheet_names or list(SHEET_EXPORTERS.keys())
    # Resolve aliases
    targets = [_ALIASES.get(t, t) for t in targets]

    for sheet_name in targets:
        if sheet_name not in SHEET_EXPORTERS:
            print(f"[skip] '{sheet_name}' not in SHEET_EXPORTERS - no parser registered")
            continue
        if sheet_name not in wb.sheetnames:
            print(f"[skip] '{sheet_name}' not found in workbook")
            continue

        parser, filename = SHEET_EXPORTERS[sheet_name]
        ws = wb[sheet_name]
        data = parser(ws)
        out_path = DATA_DIR / filename
        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        print(f"[ok]   {sheet_name:30s} -> data/{filename}")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Export Fjöld design spreadsheet tabs to JSON files in /data/."
    )
    parser.add_argument(
        "sheets",
        nargs="*",
        help=(
            "Sheet name(s) or aliases to export. "
            "If omitted, all registered sheets are exported. "
            f"Aliases: {', '.join(sorted(_ALIASES))}"
        ),
    )
    parser.add_argument(
        "--xlsx",
        type=pathlib.Path,
        default=None,
        help="Path to the workbook. Defaults to the newest Fjold_Master_Design_*.xlsx in the repo root.",
    )
    args = parser.parse_args()

    xlsx_path = args.xlsx if args.xlsx else _find_xlsx()
    print(f"Reading: {xlsx_path.name}\n")
    export_sheets(xlsx_path, args.sheets or None)
    print(f"\nDone. Files written to {DATA_DIR}/")


if __name__ == "__main__":
    main()
